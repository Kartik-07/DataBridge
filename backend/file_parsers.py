"""
File parsers — format-specific schema inference and row reading for CSV, JSON, JSONL, XLSX, Parquet.
"""
from __future__ import annotations

import csv
import io
import json
import os
import re
from datetime import date, datetime
from typing import IO, Generator

from models import ColumnInfo, FileFormat

# ── Format detection ─────────────────────────────────────────────────────────

_EXT_MAP: dict[str, FileFormat] = {
    ".json": FileFormat.JSON,
    ".jsonl": FileFormat.JSONL,
    ".ndjson": FileFormat.JSONL,
    ".csv": FileFormat.CSV,
    ".tsv": FileFormat.CSV,
    ".xlsx": FileFormat.XLSX,
    ".xls": FileFormat.XLSX,
    ".parquet": FileFormat.PARQUET,
}

_ISO_DATE_RE = re.compile(
    r"^\d{4}-\d{2}-\d{2}([ T]\d{2}:\d{2}(:\d{2}(\.\d+)?(Z|[+-]\d{2}:?\d{2})?)?)?$"
)

SCHEMA_SAMPLE_ROWS = 200  # rows sampled for type inference


def detect_format(filename: str) -> FileFormat:
    """Infer FileFormat from extension. Raises ValueError if unrecognised."""
    ext = os.path.splitext(filename)[-1].lower()
    fmt = _EXT_MAP.get(ext)
    if fmt is None:
        raise ValueError(f"Unsupported file extension: {ext!r}. Supported: {list(_EXT_MAP)}")
    return fmt


# ── Type inference helpers ───────────────────────────────────────────────────

def _infer_scalar_type(val: str) -> str:
    """Classify a single string value as a primitive type."""
    v = val.strip()
    if v == "":
        return "null"
    if v.lower() in ("true", "false"):
        return "boolean"
    try:
        int(v)
        return "integer"
    except ValueError:
        pass
    try:
        float(v)
        return "float"
    except ValueError:
        pass
    if _ISO_DATE_RE.match(v):
        return "timestamp"
    return "text"


def _merge_types(a: str, b: str) -> str:
    """Return the least-specific type that covers both a and b."""
    if a == b:
        return a
    if "null" in (a, b):
        return a if b == "null" else b
    order = ["boolean", "integer", "float", "timestamp", "text"]
    ia = order.index(a) if a in order else len(order)
    ib = order.index(b) if b in order else len(order)
    return order[max(ia, ib)]


def _python_to_db_type(py_type: str) -> str:
    """Map inferred Python type name to a generic SQL column type."""
    return {
        "boolean": "BOOLEAN",
        "integer": "BIGINT",
        "float": "DOUBLE PRECISION",
        "timestamp": "TIMESTAMP",
        "text": "TEXT",
        "null": "TEXT",
    }.get(py_type, "TEXT")


def _infer_types_from_rows(headers: list[str], sample_rows: list[list]) -> list[ColumnInfo]:
    """Build ColumnInfo list from column headers and sample rows."""
    col_types: list[str] = ["null"] * len(headers)
    col_nullable: list[bool] = [False] * len(headers)

    for row in sample_rows:
        for i, val in enumerate(row):
            if i >= len(headers):
                break
            if val is None or (isinstance(val, str) and val.strip() == ""):
                col_nullable[i] = True
            else:
                scalar = _infer_scalar_type(str(val)) if not isinstance(val, (bool, int, float, datetime, date)) else _python_native_type(val)
                col_types[i] = _merge_types(col_types[i], scalar)

    cols: list[ColumnInfo] = []
    for i, name in enumerate(headers):
        db_type = _python_to_db_type(col_types[i] if col_types[i] != "null" else "text")
        cols.append(ColumnInfo(
            name=name,
            data_type=db_type,
            is_nullable=col_nullable[i],
        ))
    return cols


def _python_native_type(val) -> str:
    if isinstance(val, bool):
        return "boolean"
    if isinstance(val, int):
        return "integer"
    if isinstance(val, float):
        return "float"
    if isinstance(val, (datetime, date)):
        return "timestamp"
    return "text"


def _infer_types_from_dicts(sample: list[dict]) -> tuple[list[str], list[ColumnInfo]]:
    """Infer schema from a list of dicts (JSON/JSONL/XLSX with headers)."""
    if not sample:
        return [], []
    # Preserve insertion-order of keys across all rows
    headers_seen: dict[str, None] = {}
    for row in sample:
        if isinstance(row, dict):
            for k in row.keys():
                headers_seen[k] = None
    headers = list(headers_seen)
    rows = [[row.get(h) for h in headers] for row in sample if isinstance(row, dict)]
    return headers, _infer_types_from_rows(headers, rows)


# ── CSV ──────────────────────────────────────────────────────────────────────

def _read_as_text(file_obj: IO[bytes]) -> io.StringIO:
    """Read the full bytes IO as a StringIO (UTF-8 with latin-1 fallback)."""
    raw = file_obj.read()
    file_obj.seek(0)
    try:
        return io.StringIO(raw.decode("utf-8"), newline=None)
    except UnicodeDecodeError:
        return io.StringIO(raw.decode("latin-1"), newline=None)


def _infer_schema_csv(file_obj: IO[bytes]) -> tuple[list[str], list[ColumnInfo]]:
    text = _read_as_text(file_obj)
    reader = csv.DictReader(text)
    sample: list[dict] = []
    for row in reader:
        sample.append(dict(row))
        if len(sample) >= SCHEMA_SAMPLE_ROWS:
            break
    file_obj.seek(0)
    return _infer_types_from_dicts(sample)


def _read_rows_csv(file_obj: IO[bytes], batch_size: int, headers: list[str]) -> Generator[list[tuple], None, None]:
    text = _read_as_text(file_obj)
    reader = csv.DictReader(text)
    batch: list[tuple] = []
    for row in reader:
        batch.append(tuple(row.get(h, None) for h in headers))
        if len(batch) >= batch_size:
            yield batch
            batch = []
    if batch:
        yield batch


# ── JSON ─────────────────────────────────────────────────────────────────────

def _infer_schema_json(file_obj: IO[bytes]) -> tuple[list[str], list[ColumnInfo]]:
    raw = file_obj.read()
    file_obj.seek(0)
    data = json.loads(raw.decode("utf-8"))
    if isinstance(data, list):
        return _infer_types_from_dicts(data[:SCHEMA_SAMPLE_ROWS])
    if isinstance(data, dict):
        return _infer_types_from_dicts([data])
    raise ValueError("JSON source must be an array of objects or a single object.")


def _read_rows_json(file_obj: IO[bytes], batch_size: int, headers: list[str]) -> Generator[list[tuple], None, None]:
    raw = file_obj.read()
    data = json.loads(raw.decode("utf-8"))
    if isinstance(data, dict):
        data = [data]
    batch: list[tuple] = []
    for row in data:
        if isinstance(row, dict):
            batch.append(tuple(row.get(h) for h in headers))
        if len(batch) >= batch_size:
            yield batch
            batch = []
    if batch:
        yield batch


# ── JSONL ─────────────────────────────────────────────────────────────────────

def _infer_schema_jsonl(file_obj: IO[bytes]) -> tuple[list[str], list[ColumnInfo]]:
    raw = file_obj.read()
    file_obj.seek(0)
    sample: list[dict] = []
    for line in raw.decode("utf-8", errors="replace").splitlines():
        line = line.strip()
        if line:
            try:
                obj = json.loads(line)
                if isinstance(obj, dict):
                    sample.append(obj)
            except json.JSONDecodeError:
                continue
        if len(sample) >= SCHEMA_SAMPLE_ROWS:
            break
    return _infer_types_from_dicts(sample)


def _read_rows_jsonl(file_obj: IO[bytes], batch_size: int, headers: list[str]) -> Generator[list[tuple], None, None]:
    raw = file_obj.read()
    batch: list[tuple] = []
    for line in raw.decode("utf-8", errors="replace").splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            obj = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(obj, dict):
            batch.append(tuple(obj.get(h) for h in headers))
            if len(batch) >= batch_size:
                yield batch
                batch = []
    if batch:
        yield batch


# ── XLSX ──────────────────────────────────────────────────────────────────────

def _infer_schema_xlsx(file_obj: IO[bytes]) -> tuple[list[str], list[ColumnInfo]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX support. Install it with: pip install openpyxl") from exc

    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers_raw = next(rows, None)
    if headers_raw is None:
        wb.close()
        return [], []
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(headers_raw)]
    sample: list[list] = []
    for row in rows:
        sample.append(list(row))
        if len(sample) >= SCHEMA_SAMPLE_ROWS:
            break
    wb.close()
    file_obj.seek(0)
    return headers, _infer_types_from_rows(headers, sample)


def _read_rows_xlsx(file_obj: IO[bytes], batch_size: int, headers: list[str]) -> Generator[list[tuple], None, None]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX support.") from exc

    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # skip header row
    batch: list[tuple] = []
    for row in rows_iter:
        batch.append(tuple(row))
        if len(batch) >= batch_size:
            yield batch
            batch = []
    if batch:
        yield batch
    wb.close()


# ── Parquet ───────────────────────────────────────────────────────────────────

def _infer_schema_parquet(file_obj: IO[bytes]) -> tuple[list[str], list[ColumnInfo]]:
    try:
        import pyarrow.parquet as pq
    except ImportError as exc:
        raise RuntimeError("pyarrow is required for Parquet support. Install it with: pip install pyarrow") from exc

    # pyarrow can read from a bytes buffer
    table = pq.read_table(file_obj)
    file_obj.seek(0)
    schema = table.schema
    cols: list[ColumnInfo] = []
    for i, field in enumerate(schema):
        db_type = _arrow_type_to_db(str(field.type))
        cols.append(ColumnInfo(
            name=field.name,
            data_type=db_type,
            is_nullable=field.nullable,
        ))
    headers = [f.name for f in schema]
    return headers, cols


def _arrow_type_to_db(arrow_type: str) -> str:
    t = arrow_type.lower()
    if "bool" in t:
        return "BOOLEAN"
    if t in ("int8", "int16", "int32", "int64", "uint8", "uint16", "uint32", "uint64"):
        return "BIGINT"
    if "float" in t or "double" in t:
        return "DOUBLE PRECISION"
    if "timestamp" in t or "date" in t:
        return "TIMESTAMP"
    if "decimal" in t:
        return "NUMERIC"
    return "TEXT"


def _read_rows_parquet(file_obj: IO[bytes], batch_size: int, headers: list[str]) -> Generator[list[tuple], None, None]:
    try:
        import pyarrow.parquet as pq
    except ImportError as exc:
        raise RuntimeError("pyarrow is required for Parquet support.") from exc

    table = pq.read_table(file_obj)
    for record_batch in table.to_batches(max_chunksize=batch_size):
        df = record_batch.to_pydict()
        n = len(next(iter(df.values()), []))
        batch = [tuple(df[h][i] for h in headers) for i in range(n)]
        if batch:
            yield batch


# ── Public API ────────────────────────────────────────────────────────────────

def infer_schema(file_obj: IO[bytes], fmt: FileFormat) -> tuple[list[str], list[ColumnInfo]]:
    """
    Read a sample from file_obj and return (headers, columns).
    file_obj is seeked back to 0 after sampling so it can be re-read.
    """
    if fmt == FileFormat.CSV:
        return _infer_schema_csv(file_obj)
    elif fmt == FileFormat.JSON:
        return _infer_schema_json(file_obj)
    elif fmt == FileFormat.JSONL:
        return _infer_schema_jsonl(file_obj)
    elif fmt == FileFormat.XLSX:
        return _infer_schema_xlsx(file_obj)
    elif fmt == FileFormat.PARQUET:
        return _infer_schema_parquet(file_obj)
    raise ValueError(f"Unsupported format: {fmt}")


def read_rows(
    file_obj: IO[bytes],
    fmt: FileFormat,
    headers: list[str],
    batch_size: int = 1000,
) -> Generator[list[tuple], None, None]:
    """
    Yield batches of row tuples in column order matching `headers`.
    file_obj should be at position 0 (or will be seeked).
    """
    if hasattr(file_obj, "seek"):
        file_obj.seek(0)

    if fmt == FileFormat.CSV:
        yield from _read_rows_csv(file_obj, batch_size, headers)
    elif fmt == FileFormat.JSON:
        yield from _read_rows_json(file_obj, batch_size, headers)
    elif fmt == FileFormat.JSONL:
        yield from _read_rows_jsonl(file_obj, batch_size, headers)
    elif fmt == FileFormat.XLSX:
        yield from _read_rows_xlsx(file_obj, batch_size, headers)
    elif fmt == FileFormat.PARQUET:
        yield from _read_rows_parquet(file_obj, batch_size, headers)
    else:
        raise ValueError(f"Unsupported format: {fmt}")
